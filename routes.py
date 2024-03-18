from flask import Flask, render_template, request, g
import random
import sqlite3
from openpyxl import load_workbook
from openpyxl import Workbook

app = Flask(__name__)

DATABASE = 'rifas.db'
MAX_REGISTROS = 100  # Máximo de registros permitidos

@app.template_filter('zfill')
def zfill(value, width):
    return str(value).zfill(width)

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.execute("CREATE TABLE IF NOT EXISTS participantes (id INTEGER PRIMARY KEY, nombre TEXT, telefono TEXT, numeros TEXT)")
        db.commit()
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def generar_numero_rifa():
    return random.randint(0, 9999)

def generar_numeros_unicos(cantidad):
    numeros_unicos = set()
    while len(numeros_unicos) < cantidad:
        numero = generar_numero_rifa()
        numeros_unicos.add(numero)
    return list(numeros_unicos)

def agregar_columna_numeros():
    with get_db() as db:
        cursor = db.cursor()
        cursor.execute("PRAGMA table_info(participantes)")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns]
        if 'numeros' not in column_names:
            cursor.execute("ALTER TABLE participantes ADD COLUMN numeros TEXT")
            db.commit()

def guardar_en_excel(participantes):
    archivo_excel = "participantes.xlsx"

    try:
        wb = load_workbook(archivo_excel)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Teléfono', 'Números'])

    ws = wb.active
    numeros_str = ", ".join(f"{num:04}" for participante in participantes for num in participante['numeros'])
    ws.append(['', '', numeros_str])  # Aquí agregamos una fila vacía para separar los registros
    for participante in participantes:
        ws.append([participante['nombre'], participante['telefono'], ""])  # Dejamos la columna de números vacía para no repetirlos

    wb.save(archivo_excel)



@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        nombre = request.form['nombre']
        telefono = request.form['telefono']
        with get_db() as db:
            cursor = db.cursor()
            agregar_columna_numeros()
            numeros_rifa = generar_numeros_unicos(5)
            cursor.execute("INSERT INTO participantes (nombre, telefono, numeros) VALUES (?, ?, ?)", (nombre, telefono, ", ".join(f"{num:04}" for num in numeros_rifa)))
            db.commit()
        guardar_en_excel([{'nombre': nombre, 'telefono': telefono, 'numeros': numeros_rifa}])
        return render_template('resultado.html', nombre=nombre, numeros_rifa=numeros_rifa)

    return render_template('index.html')


if __name__ == '__main__':
    app.run(debug=True)

