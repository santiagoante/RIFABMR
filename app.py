from flask import Flask, render_template, request, g, jsonify
import random
import sqlite3
from openpyxl import load_workbook
from openpyxl import Workbook

app = Flask(__name__)

DATABASE = 'rifas.db'
MAX_REGISTROS = 10  # Máximo de registros permitidos
N_BOLETAS = 10
MIN_DIGITOS = 0 
MAX_DIGITOS = 99



# Obtener los números asignados del archivo Excel
def obtener_numeros_asignados():
    numeros_asignados = set()
    try:
        wb = load_workbook("participantes.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                numeros = [int(num.strip()) for num in cell.value.split(",") if num.strip()]
                numeros_asignados.update(numeros)
    except FileNotFoundError:
        pass  # Archivo Excel no encontrado o no tiene números asignados
    return numeros_asignados

# Ruta para mostrar los números disponibles
@app.route('/numeros')
def mostrar_numeros_disponibles():
    numeros_asignados = obtener_numeros_asignados()
    numeros_disponibles = [num for num in range(0, MAX_REGISTROS) if num not in numeros_asignados]
    return render_template('numeros.html', numeros_disponibles=numeros_disponibles)



@app.template_filter('zfill')
def zfill(value, width):
    return str(value).zfill(width)

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.execute("CREATE TABLE IF NOT EXISTS participantes (id INTEGER PRIMARY KEY, nombre TEXT, telefono TEXT, numeros TEXT)")
        db.commit()
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def generar_numero_rifa():
    return random.randint(MIN_DIGITOS, MAX_DIGITOS)

def obtener_numeros_asignados():
    with get_db() as db:
        cursor = db.cursor()
        cursor.execute("SELECT numeros FROM participantes")
        numeros_asignados = cursor.fetchall()
    # Convertir la cadena de números separados por comas en una lista de enteros
    numeros = []
    for row in numeros_asignados:
        numeros.extend(map(int, row[0].split(', ')))
    return numeros

def numeros_unicos_disponibles(cantidad):
    numeros_asignados = obtener_numeros_asignados()
    numeros_disponibles = set(range(100)) - set(numeros_asignados)
    return len(numeros_disponibles) >= cantidad


def generar_numeros_unicos(cantidad):
    numeros_asignados = obtener_numeros_asignados()
    numeros_unicos = set()
    while len(numeros_unicos) < cantidad:
        numero = generar_numero_rifa()
        if numero not in numeros_asignados: # Verificar si el número no está asignado 
            numeros_unicos.add(numero)
    return list(numeros_unicos)



def obtener_numero_registros_db():
    with get_db() as db:
        cursor = db.cursor()
        cursor.execute("SELECT COUNT(*) FROM participantes")
        numero_registros = cursor.fetchone()[0]
    return numero_registros

def obtener_numero_registros_excel():
    archivo_excel = "participantes.xlsx"
    try:
        df = load_workbook(archivo_excel)
        ws = df.active
        return ws.max_row - 1  # Restamos 1 para excluir la fila de encabezado
    except FileNotFoundError:
        return 0

def obtener_porcentaje_registrados():
    numero_registrados_db = obtener_numero_registros_db()
    numero_registrados_excel = obtener_numero_registros_excel()
    total_registros = numero_registrados_db + numero_registrados_excel
    if total_registros == 0:
        return 0
    else:
        return (total_registros / 10000) * 100  # Calcula el porcentaje en relación con 10000 registros


def guardar_en_excel(participantes):
    archivo_excel = "participantes.xlsx"

    try:
        wb = load_workbook(archivo_excel)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Teléfono', 'Números'])

    ws = wb.active
    for participante in participantes:
        ws.append([participante['nombre'], participante['telefono'], ", ".join(f"{num:04}" for num in participante['numeros'])])

    wb.save(archivo_excel)

def agregar_columna_numeros():
    with get_db() as db:
        cursor = db.cursor()
        cursor.execute("PRAGMA table_info(participantes)")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns]
        if 'numeros' not in column_names:
            cursor.execute("ALTER TABLE participantes ADD COLUMN numeros TEXT")
            db.commit()

def obtener_numeros_asignados():
    with get_db() as db:
        cursor = db.cursor()
        cursor.execute("SELECT numeros FROM participantes")
        numeros_asignados = cursor.fetchall()
    # Convertir la cadena de números separados por comas en una lista de enteros
    numeros = []
    for row in numeros_asignados:
        numeros.extend(map(int, row[0].split(', ')))  # Aquí se separa por ', '
    return numeros

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        nombre = request.form['nombre']
        telefono = request.form['telefono']
        
        if not numeros_unicos_disponibles(N_BOLETAS):
            # Si no hay más números únicos disponibles, muestra una pantalla de éxito y reinicia los números únicos
            mensaje = "¡Éxito! Todos los números únicos han sido asignados. Se reiniciaron los números para nuevos registros."
            reiniciar_numeros_unicos()
            return render_template('exito.html', mensaje=mensaje)
        
        with get_db() as db:
            cursor = db.cursor()
            agregar_columna_numeros()
            # Genera 10 números aleatorios de 4 dígitos
            numeros_rifa = generar_numeros_unicos(N_BOLETAS)
            cursor.execute("INSERT INTO participantes (nombre, telefono, numeros) VALUES (?, ?, ?)", (nombre, telefono, ", ".join(f"{num:04}" for num in numeros_rifa)))
            db.commit()
            
            # Calcular el porcentaje de registros completados
            porcentaje_registrados = obtener_porcentaje_registrados()
        

            # Obtener el número de registros en el archivo Excel
            numero_registrados_excel = obtener_numero_registros_excel()

            # Calcular el número total de registros
            total_registros = numero_registrados_excel

            # Obtener el porcentaje de registros en comparación con el máximo de registros permitidos
            porcentaje_registrados = (total_registros / MAX_REGISTROS) * 100 if MAX_REGISTROS > 0 else 0
            
        guardar_en_excel([{'nombre': nombre, 'telefono': telefono, 'numeros': numeros_rifa}])
        
        return render_template('resultado.html', nombre=nombre, numeros_rifa=numeros_rifa, porcentaje_registrados=porcentaje_registrados)
    else:
        # Obtener el número de registros en el archivo Excel
        numero_registrados_excel = obtener_numero_registros_excel()

        # Calcular el número total de registros
        total_registros =  numero_registrados_excel

        # Obtener el porcentaje de registros en comparación con el máximo de registros permitidos
        porcentaje_registrados = (total_registros / MAX_REGISTROS) * 100 if MAX_REGISTROS > 0 else 0
        
        return render_template('index.html', porcentaje_registrados=porcentaje_registrados)

def reiniciar_numeros_unicos():
    with get_db() as db:
        cursor = db.cursor()
        cursor.execute("DELETE FROM participantes")  # Eliminar todos los registros de la tabla 'participantes'
        db.commit()

def reiniciar_numeros_unicos():
    with get_db() as db:
        cursor = db.cursor()
        cursor.execute("DELETE FROM participantes")  # Eliminar todos los registros de la tabla 'participantes'
        db.commit()



@app.route('/participantes')
def ver_participantes():
    with get_db() as db:
        cursor = db.cursor()
        cursor.execute("SELECT nombre, telefono, numeros FROM participantes")
        participantes = cursor.fetchall()
    return render_template('participantes.html', participantes=participantes)

@app.route('/numeros_disponibles')
def numeros_disponibles():
    numeros_asignados = obtener_numeros_asignados()
    numeros_disponibles = [numero for numero in range(MAX_DIGITOS + 1) if numero not in numeros_asignados]
    return jsonify({'numeros_disponibles': numeros_disponibles})

@app.route('/num_registros', methods=['GET'])
def obtener_numero_registros():
    numero_registros_excel = obtener_numero_registros_excel()
    total_registros =  numero_registros_excel
    return jsonify({'num_registros': total_registros})

def calcular_porcentaje(valor_actual, valor_total):
    if valor_total == 0:
        return 0
    else:
        return (valor_actual / valor_total) * 100

@app.route('/exito')
def exito():
    # Esta es la ruta para la página de éxito
    return render_template('exito.html')

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000)